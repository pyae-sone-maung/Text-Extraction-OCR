from openpyxl import load_workbook
import  os

upload_dri = os.path.join(os.path.dirname(__file__), "..", "..", "assets")
file_path = os.path.join(upload_dri, "excel_documents", 'sample.xlsx')

extract_text = []

workbook = load_workbook(file_path)
for sheet in workbook.worksheets:
    for row in sheet.iter_rows():
        row_data = []
        for cell in row:
            if cell.value is not None:
                if cell.data_type != 'f':
                    cell_value = str(cell.value) if cell.value is not None else ""
                    row_data.append(cell_value)
        extract_text.append(" ".join(row_data))
    extract_text.append(" ")

extract_text = "\n".join(extract_text)
print(extract_text)